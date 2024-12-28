import os

from openpyxl import Workbook

from interface.widgets.message import message
from managers.DAO_classes import customer_DAO, project_DAO, stuff_DAO, tasks_DAO


class report_manager:
    @staticmethod
    def make_projects_report():
        """Створює звіт"""
        report = report_manager.get_projects_info()
        report_manager.create_project_report(report, "projects_report.xlsx")

    @staticmethod
    def get_projects_info():
        """Повертає дані у вигляді однієї структури"""
        report = []

        projects = project_DAO.get_all_projects()

        for project in projects:

            customer = customer_DAO.get_customer_by_id(project.customer_id)
            tasks = tasks_DAO.get_tasks_by_project_id(project.id)

            report.append(
                {
                    "Project_name": project.name,
                    "Customer": {
                        "Name": customer.name,
                        "Surname": customer.surname,
                        "Phone": customer.phone,
                        "Address": customer.address,
                        "Mail": customer.mail,
                    },
                    "Paid": project.cost,
                    "Cost": project.paid,
                    "Tasks": [
                        {
                            "deadline": task.deadline,
                            "is_done": task.is_done,
                            "is_checked": task.is_checked,
                            "Comment": task.comment,
                        }
                        for task in tasks
                    ],
                }
            )

        return report

    @staticmethod
    def create_project_report(report, filename):
        """Створює EXCEL файл"""
        wb = Workbook()

        for project in report:
            project_name = project["Project_name"]
            ws = wb.create_sheet(title=project_name)

            customer = project["Customer"]
            ws.append(
                [
                    f"Customer Name: {customer['Name']} {customer['Surname']}",
                    f"Phone: {customer['Phone']}",
                    f"Address: {customer['Address']}",
                    f"Email: {customer['Mail']}",
                ]
            )

            ws.append(
                [
                    f"Paid: {project['Paid']}",
                    f"Cost: {project['Cost']}",
                ]
            )

            ws.append([])

            ws.append(["deadline", "is_done", "is_checked", "Comment"])

            for task in project["Tasks"]:
                task_row = [
                    task["deadline"],
                    task["is_done"],
                    task["is_checked"],
                    task["Comment"],
                ]
                ws.append(task_row)

            column_widths = {
                "A": 36,
                "B": 22,
                "C": 30,
                "D": 46,
            }

            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

        del wb["Sheet"]

        if not os.path.exists("reports"):
            os.makedirs("reports")
        wb.save("reports/" + filename)

        message.show_message("Звіт сформовано", "Перевірте папку reports")

    @staticmethod
    def make_stuff_report(stuff_list):
        report = report_manager.get_stuff_info(stuff_list)
        report_manager.create_stuff_report(report, "stuff_report.xlsx")

    @staticmethod
    def get_stuff_info(stuff_list):
        """Повертає інформацію про працівників та їхні таски"""
        if not stuff_list:
            message.show_message("Помилка", "Група не може бути порожньою")
            return []

        report = []
        for stuff in stuff_list:
            staff_id = stuff_DAO.get_staff_id_by_fullname(stuff)
            if staff_id:
                staff_details = stuff_DAO.get_stuff_by_id(staff_id)
                tasks = tasks_DAO.get_tasks_by_staff_id_with_names(staff_id)

                report.append(
                    {
                        "Name": staff_details.name,
                        "Surname": staff_details.surname,
                        "Role": staff_details.role,
                        "Phone": staff_details.phone,
                        "Address": staff_details.address,
                        "Mail": staff_details.mail,
                        "Date_of_Birth": staff_details.date_of_birth,
                        "Tasks": tasks,
                    }
                )
        return report

    @staticmethod
    def create_stuff_report(report, filename):
        """Створює Excel-файл для звіту працівників"""
        if not report:
            return

        wb = Workbook()

        for staff in report:
            sheet_name = f"{staff['Name']} {staff['Surname']}"
            ws = wb.create_sheet(title=sheet_name)

            # Додавання основної інформації про працівника
            ws.append(
                [
                    f"Name: {staff['Name']}",
                    f"Surname: {staff['Surname']}",
                    f"Role: {staff['Role']}",
                ]
            )
            ws.append(
                [
                    f"Phone: {staff['Phone']}",
                    f"Address: {staff['Address']}",
                    f"Mail: {staff['Mail']}",
                ]
            )
            ws.append([f"Date of Birth: {staff['Date_of_Birth']}"])
            ws.append([])

            # Заголовки для таблиці тасків
            ws.append(["Deadline", "Is Done", "Is Checked", "Comment"])

            # Додавання інформації про таски
            for task in staff["Tasks"]:
                ws.append(
                    [
                        task["deadline"],
                        task["is_done"],
                        task["is_checked"],
                        task["Comment"],
                    ]
                )

            # Встановлення ширини колонок
            column_widths = {"A": 25, "B": 25, "C": 25, "D": 30}
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

        # Видалення стандартного аркуша, якщо він є
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Збереження файлу
        if not os.path.exists("reports"):
            os.makedirs("reports")

        wb.save(f"reports/{filename}")
        message.show_message("Звіт сформовано", "Перевірте папку reports")
